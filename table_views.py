from PyQt5.QtWidgets import QWidget, QFrame , QMenu, QDialogButtonBox, QVBoxLayout, QLabel, QTableWidget, QTextEdit, QTableWidgetItem, QHeaderView, QHBoxLayout, QLineEdit, QPushButton, QFileDialog, QMessageBox, QInputDialog, QDateEdit, QSpacerItem, QDialog, QSizePolicy
from PyQt5.QtGui import QFont, QPixmap, QColor, QCursor
from PyQt5.QtCore import Qt, QDate
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os
import subprocess

from dialogs import AddConsumerDialog, ProjectInputDialog, AddTukangDialog, AddMaterialDialog, AddSalesProjectDialog, AddTukangProjectDialog
from database import DatabaseManager, COLUMN_MAPPINGS
from error_handling import setup_error_handling

def format_backup_name(backup_name, table_type, person_name=None):
    parts = backup_name.split('_')
    year = parts[-4]
    month = parts[-3]
    day = parts[-2]
    count = parts[-1]
    month_name = datetime(int(year), int(month), 1).strftime('%B')

    if table_type == 'consumer':
        return f"Konsumen {year} {month_name} {day} ({count})"
    elif table_type in ['sales', 'tukang']:
        return f"{person_name} {year} {month_name} {day} ({count})"
    else:
        return f"{year} {month_name} {day} ({count})"

def format_rupiah(value):
    try:
        numeric_value = float(value)
        return f"Rp {numeric_value:,.0f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except ValueError:
        return value  # Return original value if conversion fails

class TableWidget(QWidget):
    def __init__(self, table_name, parent=None):
        super().__init__(parent)
        setup_error_handling()
        self.table_name = table_name
        self.db = DatabaseManager()
        self.is_viewing_history = False
        self.current_book_name = None
        self.user_id = None
        self.setup_ui()

        self.table.cellDoubleClicked.connect(self.show_full_note)

    def setup_ui(self):
        self.layout = QVBoxLayout(self)
        
        self.title_label = QLabel()
        self.title_label.setFont(QFont("Arial", 16, QFont.Bold))
        self.layout.addWidget(self.title_label)

        self.setup_search()

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.layout.addWidget(self.table)

        self.setup_buttons()

    def setup_search(self):
        self.search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Search...")
        self.search_input.textChanged.connect(self.filter_table)
        self.search_layout.addWidget(self.search_input)
        self.layout.addLayout(self.search_layout)

    def filter_table(self):
        search_text = self.search_input.text().lower()
        for row in range(self.table.rowCount()):
            match = False
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                if item and search_text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)
    
    def show_full_note(self, row, column):
        item = self.table.item(row, column)
        if item:
            header_item = self.table.horizontalHeaderItem(column)
            column_name = header_item.text() if header_item else f"Column {column}"
            note = item.text()
    
            dialog = QDialog(self)
            dialog.setWindowTitle(f"{column_name} Lengkap")
            layout = QVBoxLayout(dialog)
    
            text_edit = QTextEdit(dialog)
            text_edit.setPlainText(note)
            text_edit.setReadOnly(True)
            text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
            text_edit.setMinimumSize(600, 300)
            layout.addWidget(text_edit)
    
            close_button = QPushButton("Tutup", dialog)
            close_button.clicked.connect(dialog.close)
            layout.addWidget(close_button)
    
            dialog.setLayout(layout)
            dialog.exec_()


    def setup_buttons(self):
        self.button_layout = QHBoxLayout()
        
        self.add_button = QPushButton("Tambah")
        self.add_button.clicked.connect(self.open_add_dialog)
        self.button_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Edit")
        self.edit_button.clicked.connect(self.open_edit_dialog)
        self.button_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Hapus")
        self.delete_button.clicked.connect(self.delete_selected_row)
        self.button_layout.addWidget(self.delete_button)

        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.button_layout.addWidget(self.export_button)

        if not isinstance(self, MaterialTable):
            self.view_history_button = QPushButton("Lihat Riwayat")
            self.view_history_button.clicked.connect(self.view_history)
            self.button_layout.addWidget(self.view_history_button)
            
            self.close_book_button = QPushButton("Tutup Buku")
            self.close_book_button.clicked.connect(self.close_book)
            self.button_layout.addWidget(self.close_book_button)

        self.layout.addLayout(self.button_layout)

    def set_user_id(self, user_id):
        self.user_id = user_id
        self.load_data()
    
    def add_to_closed_book(self, data):
        header_labels = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        data_dict = dict(zip(header_labels, data))
        self.db.add_to_closed_book(self.current_book_name, data_dict)

    def update_in_closed_book(self, record_id, data):
        header_labels = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        data_dict = dict(zip(header_labels, data))
        self.db.update_in_closed_book(self.current_book_name, record_id, data_dict)

    def filter_table(self):
        search_text = self.search_input.text().lower()
        for row in range(self.table.rowCount()):
            match = False
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                if item and search_text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)

    def close_book(self):
        # Check if the table is empty
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, 'Tutup Buku', 'Tidak dapat menutup buku karena data kosong.')
            return

        reply = QMessageBox.question(self, 'Tutup Buku', 'Anda yakin ingin menutup buku? Ini akan membuat backup data saat ini dan menghapus semua data dari tabel.',
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            backup_name = self.db.close_book(self.table_name)
            self.load_data()  # Reload the (now empty) table
            QMessageBox.information(self, "Tutup Buku", f"Buku telah ditutup. Data lama telah dibackup ke {backup_name} dan tabel telah direset.")

    def view_history(self):
        closed_books = self.db.get_closed_books(self.table_name)
        if closed_books:
            book, ok = QInputDialog.getItem(self, "Pilih Riwayat", f"Riwayat {self.title_label.text()}:", closed_books, 0, False)
            if ok:
                data = self.db.load_closed_book(book)
                self.display_history(data, book)
        else:
            QMessageBox.information(self, "Tidak Ada Data", "Tidak ada riwayat tersimpan untuk tabel ini.")

    def display_history(self, data, book_name):
        self.table.setRowCount(0)  # Clear the current table
        self.title_label.setText(f"Riwayat {self.title_label.text().split(' - ')[0]} - {book_name}")
    
        for row_data in data:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            for column, item in enumerate(row_data[1:]):  # Exclude id
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))
    
        self.close_book_button.hide()
        self.view_history_button.hide()
    
        # Perbaikan pada baris ini
        self.return_button = QPushButton("Kembali ke Data Saat Ini")
        self.return_button.clicked.connect(self.return_to_current_data)
        self.button_layout.addWidget(self.return_button)

        self.current_book_name = book_name
        self.is_viewing_history = True

    def return_to_current_data(self):
        self.title_label.setText(self.title_label.text().split(" - ")[0])  # Remove the "Riwayat" part
        self.is_viewing_history = False
        self.current_book_name = None
        self.load_data()
        self.close_book_button.show()
        self.view_history_button.show()
        if self.return_button:
            self.return_button.setParent(None)  # Remove the return button
            self.return_button = None

    def load_data(self):
        self.table.setRowCount(0)  # Clear existing data
        if self.is_viewing_history:
            data = self.db.load_closed_book(self.current_book_name)
        else:
            current_date = datetime.now()
            data = getattr(self.db, f"get_{self.table_name}")(year=current_date.year, month=current_date.month)
        
        for row_data in data:
            self.add_row(row_data[1:])  # Exclude id

    def open_add_dialog(self):
        dialog = self.get_add_dialog()
        if dialog.exec_():
            data = dialog.get_data()
            if self.is_viewing_history:
                header_labels = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
                data_dict = dict(zip(header_labels, data))
                new_id = self.db.add_to_closed_book(self.current_book_name, data_dict)
                self.add_row(data, new_id)
            else:
                insert_method = getattr(self.db, f"insert_{self.table_name[:-1]}")
                current_date = datetime.now()
                new_id = insert_method(data, year=current_date.year, month=current_date.month)
                self.add_row(data, new_id)
            
    def open_edit_dialog(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin diedit terlebih dahulu.")
            return

        selected_row = selected_items[0].row()
        dialog = self.get_edit_dialog()

        initial_data = []
        for col in range(self.table.columnCount()):
            item = self.table.item(selected_row, col)
            if item is not None:
                initial_data.append(item.text())
            else:
                initial_data.append("")

        dialog.load_data(initial_data)

        if dialog.exec_():
            data = dialog.get_data()
            for col, item_text in enumerate(data):
                self.table.setItem(selected_row, col, QTableWidgetItem(str(item_text) if item_text is not None else ""))
    
            if self.is_viewing_history:
                record_id = self.db.load_closed_book(self.current_book_name)[selected_row][0]
                header_labels = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
                data_dict = dict(zip(header_labels, data))
                self.db.update_in_closed_book(self.current_book_name, record_id, data_dict)
            else:
                table_data = getattr(self.db, f"get_{self.table_name}")()
                record_id = table_data[selected_row][0]
                update_method = getattr(self.db, f"update_{self.table_name[:-1]}", None)
                if update_method:
                    update_method(record_id, data)

    def delete_selected_row(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin dihapus terlebih dahulu.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus data ini?',
                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            selected_row = selected_items[0].row()

            if self.is_viewing_history:
                closed_book_data = self.db.load_closed_book(self.current_book_name)
                if selected_row < len(closed_book_data):
                    record_id = closed_book_data[selected_row][0]
                    self.db.delete_from_closed_book(self.current_book_name, record_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                    return
            else:
                if self.table_name == "materials_usage":
                    record_id = self.db.get_material_usage(self.current_project_id)[selected_row][0]
                else:
                    data = getattr(self.db, f"get_{self.table_name}")()
                    record_id = data[selected_row][0]
                self.db.delete_record(self.table_name, record_id)

        self.table.removeRow(selected_row)
        QMessageBox.information(self, "Hapus Data", "Data berhasil dihapus.")

    def export_to_excel(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if file_name:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Write headers
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount()) if not self.table.isColumnHidden(i)]
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Write data
            for row in range(self.table.rowCount()):
                col_index = 1
                for column in range(self.table.columnCount()):
                    if not self.table.isColumnHidden(column):
                        item = self.table.item(row, column)
                        if item is not None:
                            cell = sheet.cell(row=row+2, column=col_index, value=item.text())
                            if headers[col_index-1] == "Keterangan":
                                cell.alignment = Alignment(wrapText=True, vertical='top')
                                cell = self.format_keterangan_cell(cell)
                        else:
                            sheet.cell(row=row+2, column=col_index, value='')
                        col_index += 1

            sheet.append([""] * sheet.max_column)            
            self.add_additional_info(sheet)
            
            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_name)
            QMessageBox.information(self, "Export Successful", f"Data has been exported to {file_name}")

    def format_keterangan_cell(self, cell, words_per_line=10):
        if cell.value:
            words = cell.value.split()
            lines = []
            current_line = []
            for word in words:
                current_line.append(word)
                if len(current_line) == words_per_line:
                    lines.append(' '.join(current_line))
                    current_line = []
            if current_line:
                lines.append(' '.join(current_line))
            cell.value = '\n'.join(lines)
        return cell

    def add_additional_info(self, sheet):
        # This method can be overridden in subclasses to add table-specific information
        pass

    def get_edit_dialog(self):
        # This method should be implemented in child classes
        pass

    def add_row(self, data, new_id=None):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        for column, item in enumerate(data):
            self.table.setItem(row_position, column, QTableWidgetItem(str(item)))

        # Make sure the ID column exists
        if 'ID' not in [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]:
            self.table.insertColumn(self.table.columnCount())
            self.table.setHorizontalHeaderItem(self.table.columnCount() - 1, QTableWidgetItem("ID"))
            self.table.hideColumn(self.table.columnCount() - 1)

        # Set the ID in the last column if provided
        if new_id is not None:
            self.table.setItem(row_position, self.table.columnCount() - 1, QTableWidgetItem(str(new_id)))
        elif len(data) > 0:  # Use the first column as ID if new_id is not provided
            self.table.setItem(row_position, self.table.columnCount() - 1, QTableWidgetItem(str(data[0])))

class ConsumerTable(TableWidget):
    def __init__(self, parent=None):
        super().__init__("consumers", parent)
        setup_error_handling()
        self.title_label.setText("Daftar Konsumen")
        self.setup_table()

    def setup_table(self):
        self.table.setColumnCount(len(COLUMN_MAPPINGS[self.table_name]))
        self.table.setHorizontalHeaderLabels(list(COLUMN_MAPPINGS[self.table_name].keys()))

    def load_data(self):
        self.table.setRowCount(0)  # Clear existing data
        if self.user_id:
            current_date = datetime.now()
            if self.is_viewing_history:
                data = self.db.load_closed_book(self.current_book_name)
            else:
                data = self.db.get_consumers(year=current_date.year, month=current_date.month, user_id=self.user_id)
            for row_data in data:
                formatted_data = list(row_data[1:])  # Exclude id
                formatted_data[5] = self.format_currency(formatted_data[5])  # Format total proyek
                self.add_row(formatted_data)
    
    def add_row(self, data, new_id=None):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        for column, item in enumerate(data):
            if column == 5:  # Kolom Total Proyek
                formatted_value = self.format_currency(item)
                self.table.setItem(row_position, column, QTableWidgetItem(formatted_value))
            else:
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))

        # Make sure the ID column exists
        if 'ID' not in [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]:
            self.table.insertColumn(self.table.columnCount())
            self.table.setHorizontalHeaderItem(self.table.columnCount() - 1, QTableWidgetItem("ID"))
            self.table.hideColumn(self.table.columnCount() - 1)

        # Set the ID in the last column if provided
        if new_id is not None:
            self.table.setItem(row_position, self.table.columnCount() - 1, QTableWidgetItem(str(new_id)))
        elif len(data) > 0:  # Use the first column as ID if new_id is not provided
            self.table.setItem(row_position, self.table.columnCount() - 1, QTableWidgetItem(str(data[0])))

    def open_add_dialog(self):
        dialog = AddConsumerDialog(self)
        if dialog.exec_():
            data = dialog.get_data()
            current_date = QDate.currentDate().toString("dd/MM/yyyy")  # Add current date as the first item
            print(f"Data: {data}")

            if not self.validate_numeric_input(data[5]):  # Adjust index for Total Proyek
                QMessageBox.warning(self, "Error", "Total proyek harus berupa angka tanpa huruf atau karakter khusus.")
                return

            try:
                data[5] = self.format_currency(self.parse_currency(data[5]))  # Adjust index for Total Proyek
            except (IndexError, ValueError) as e:
                QMessageBox.warning(self, "Error", f"Total proyek tidak valid: {str(e)}")
                return

            if self.is_viewing_history:
                self.db.add_to_closed_book(self.current_book_name, dict(zip(COLUMN_MAPPINGS['consumers'].keys(), data)))
                self.load_data()
            else:
                current_date = datetime.now()
                new_id = self.db.insert_consumer(data, year=current_date.year, month=current_date.month, user_id=self.user_id)
                self.add_row((new_id,) + tuple(data))
                self.load_data()

            QMessageBox.information(self, "Sukses", "Data konsumen berhasil ditambahkan.")

    def open_edit_dialog(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin diedit terlebih dahulu.")
            return

        selected_row = selected_items[0].row()
        dialog = AddConsumerDialog(self)

        initial_data = []
        for col in range(self.table.columnCount()):
            item = self.table.item(selected_row, col)
            if item is not None:
                if col == 5:  # Kolom Total Proyek
                    initial_data.append(str(self.parse_currency(item.text())))
                else:
                    initial_data.append(item.text())
            else:
                initial_data.append("")
        
        print(f"Initial data: {initial_data}")
        dialog.load_data(initial_data)

        if dialog.exec_():
            data = dialog.get_data()
            print(f"Data: {data}")
    
            # Validasi dan format total proyek
            if not self.validate_numeric_input(data[5]):
                QMessageBox.warning(self, "Error", "Total proyek harus berupa angka tanpa huruf atau karakter khusus.")
                return

            try:
                data[5] = self.format_currency(self.parse_currency(data[5]))
            except (IndexError, ValueError) as e:
                QMessageBox.warning(self, "Error", f"Total proyek tidak valid: {str(e)}")
                return

            for col, item_text in enumerate(data):
                self.table.setItem(selected_row, col, QTableWidgetItem(str(item_text) if item_text is not None else ""))

            if self.is_viewing_history:
                record_id = self.db.load_closed_book(self.current_book_name)[selected_row][0]
                self.db.update_in_closed_book(self.current_book_name, record_id, dict(zip(COLUMN_MAPPINGS['consumers'].keys(), data)))
                self.load_data()  # Reload history data to reflect changes
            else:
                consumers = self.db.get_consumers(user_id=self.user_id)
                record_id = consumers[selected_row][0]
                self.db.update_consumer(record_id, data, self.user_id)
                self.load_data()  # Reload current data to reflect changes

            QMessageBox.information(self, "Sukses", "Data konsumen berhasil diedit.")
    
    def validate_numeric_input(self, value):
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        return cleaned_value.isdigit()
    
    def format_currency(self, value):
        try:
            if isinstance(value, str):
                value = self.parse_currency(value)
            rounded_value = int(value)
            return f"Rp {rounded_value:,}".replace(',', '.')
        except ValueError:
            return str(value)

    def parse_currency(self, value):
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        try:
            return int(float(cleaned_value))
        except ValueError:
            return 0

    def delete_selected_row(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin dihapus terlebih dahulu.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus data ini?',
                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            selected_row = selected_items[0].row()
            print(f"Selected row: {selected_row}")

            if self.is_viewing_history:
                closed_book_data = self.db.load_closed_book(self.current_book_name)
                print(f"Closed book data: {closed_book_data}")
                if selected_row < len(closed_book_data):
                    record_id = closed_book_data[selected_row][0]
                    self.db.delete_from_closed_book(self.current_book_name, record_id)
                else:
                    error_message = "Tidak dapat menemukan data yang akan dihapus."
                    QMessageBox.warning(self, "Error", error_message)
                    print(f"Error: {error_message}")
                    return
            else:
                consumers = self.db.get_consumers(user_id=self.user_id)
                print(f"Consumers: {consumers}")
                if selected_row < len(consumers):
                    record_id = consumers[selected_row][0]
                    self.db.delete_record(self.table_name, record_id, self.user_id)
                else:
                    error_message = "Tidak dapat menemukan data yang akan dihapus."
                    QMessageBox.warning(self, "Error", error_message)
                    print(f"Error: {error_message}")
                    return

        self.table.removeRow(selected_row)
        QMessageBox.information(self, "Hapus Data", "Data berhasil dihapus.")

    def close_book(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, 'Tutup Buku', 'Tidak dapat menutup buku karena data kosong.')
            return

        reply = QMessageBox.question(self, 'Tutup Buku', 'Anda yakin ingin menutup buku? Ini akan membuat backup data saat ini dan menghapus semua data dari tabel.',
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            backup_name = self.db.close_book(self.table_name, self.user_id)
            self.load_data()  # Reload the (now empty) table
            QMessageBox.information(self, "Tutup Buku", f"Buku telah ditutup. Data lama telah dibackup ke {backup_name} dan tabel telah direset.")

    def view_history(self):
        closed_books = self.db.get_closed_books(self.table_name, self.user_id)
        if closed_books:
            formatted_names = {self.format_backup_name(book): book for book in closed_books}
        
            formatted_name, ok = QInputDialog.getItem(
                self, 
                "Pilih Riwayat", 
                "Riwayat Konsumen:", 
                sorted(formatted_names.keys(), reverse=True),
                0, 
                False
            )
        
            if ok:
                original_name = formatted_names[formatted_name]
                data = self.db.load_closed_book(original_name)
                self.display_history(data, formatted_name, original_name)
        else:
            QMessageBox.information(self, "Tidak Ada Data", "Tidak ada riwayat tersimpan untuk konsumen.")

    def format_backup_name(self, backup_name):
        parts = backup_name.split('_')
        year = parts[-4]
        month = parts[-3]
        day = parts[-2]
        count = parts[-1]
        
        month_name = datetime(int(year), int(month), 1).strftime('%B')
        
        return f"Konsumen {year} {month_name} {day} ({count})"

    def display_history(self, data, book_name, original_name):
        self.table.setRowCount(0)  # Clear the current table
        self.title_label.setText(f"Riwayat {self.title_label.text().split(' - ')[0]} - {book_name}")
    
        for row_data in data:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            for column, item in enumerate(row_data[1:]):  # Exclude id
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))
    
    
        self.close_book_button.hide()
        self.view_history_button.hide()
    
        self.return_button = QPushButton("Kembali ke Data Saat Ini")
        self.return_button.clicked.connect(self.return_to_current_data)
        self.button_layout.addWidget(self.return_button)

        self.current_book_name = original_name
        self.is_viewing_history = True

    def return_to_current_data(self):
        self.title_label.setText("Daftar Konsumen")
        self.is_viewing_history = False
        self.current_book_name = None
        self.load_data()
        self.close_book_button.show()
        self.view_history_button.show()
        self.add_button.setEnabled(True)
        self.edit_button.setEnabled(True)
        self.delete_button.setEnabled(True)
        self.return_button.setParent(None)
        self.return_button = None

class SalesTable(TableWidget):
    def __init__(self, parent=None):
        super().__init__("sales_projects", parent)
        setup_error_handling()
        self.title_label.setText("Daftar Proyek Sales")
        self.current_sales_id = None
        self.current_sales_name = ""
        self.total_project_column_index = 4
        self.commission_column_index = 5
        self.kb_column_index = 6
        self.setup_table()
        self.setup_sales_buttons()
        self.setup_total_commission_label()
        self.setup_view_photo_button()


    def set_user_id(self, user_id):
        super().set_user_id(user_id)
        self.current_sales_id = None
        self.current_sales_name = ""
        self.load_data()

    def setup_table(self):
        column_names = ["ID"] + list(COLUMN_MAPPINGS[self.table_name].keys())
        self.table.setColumnCount(len(column_names))
        self.table.setHorizontalHeaderLabels(column_names)
        self.table.hideColumn(0)  # Hide ID column
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
    
    def setup_total_commission_label(self):
        self.total_commission_label = QLabel("Total Komisi: Rp 0")
        self.total_commission_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        self.total_commission_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.layout.addWidget(self.total_commission_label)

        self.total_kb_label = QLabel("Total KB: Rp 0")
        self.total_kb_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        self.total_kb_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.layout.addWidget(self.total_kb_label)

    def setup_sales_buttons(self):

        self.select_sales_button = QPushButton("Pilih Sales")
        self.select_sales_button.clicked.connect(self.select_sales)
        self.button_layout.addWidget(self.select_sales_button)

        self.setting_button = QPushButton("Setting")
        self.setting_button.clicked.connect(self.open_settings)
        self.button_layout.addWidget(self.setting_button)
        self.setting_button.hide()

    def open_settings(self):
        menu = QDialog(self)
        menu.setWindowTitle("Setting Sales")
        menu.setMinimumWidth(400)
        layout = QVBoxLayout(menu)

        new_sales_button = QPushButton("Sales Baru")
        new_sales_button.clicked.connect(lambda: self.create_new_sales(menu))
        layout.addWidget(new_sales_button)

        edit_sales_button = QPushButton("Edit Sales")
        edit_sales_button.clicked.connect(lambda: self.edit_sales(menu))
        layout.addWidget(edit_sales_button)

        delete_sales_button = QPushButton("Hapus Sales")
        delete_sales_button.clicked.connect(lambda: self.delete_sales(menu))
        layout.addWidget(delete_sales_button)

        # Disable edit and delete buttons if no sales is selected
        edit_sales_button.setEnabled(self.current_sales_id is not None)
        delete_sales_button.setEnabled(self.current_sales_id is not None)

        menu.setLayout(layout)
        menu.exec_()
    
    def view_photo(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()

        if self.is_viewing_history:
            closed_book_data = self.db.load_closed_book(self.current_book_name)
            if selected_row < len(closed_book_data):
                project_id = closed_book_data[selected_row][0]
                photo_path = closed_book_data[selected_row][-1]  # Assuming photo_path is the last column
            else:
                QMessageBox.warning(self, "Error", "Tidak dapat menemukan data proyek.")
                return
        else:
            project_id = self.table.item(selected_row, 0).text()  # Assuming ID is in the first column
            photo_path = self.db.get_sales_project_photo(project_id, self.user_id)

        if photo_path and os.path.exists(photo_path):
            dialog = PhotoViewerDialog(self, photo_path)
            dialog.exec_()
        else:
            QMessageBox.information(self, "Tidak Ada Gambar", "Tidak ada gambar tersedia untuk proyek ini.")

    def setup_view_photo_button(self):
        self.view_photo_button = QPushButton("Lihat Gambar")
        self.view_photo_button.clicked.connect(self.view_photo)
        self.view_photo_button.hide()  # Sembunyikan tombol saat inisialisasi
        self.button_layout.addWidget(self.view_photo_button)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)


    def on_selection_changed(self):
        selected_items = self.table.selectedItems()
        if selected_items:
            self.view_photo_button.show()
        else:
            self.view_photo_button.hide()


    def create_new_sales(self, menu):
        sales_name, ok = QInputDialog.getText(self, "Sales Baru", "Nama Sales:")
        if ok and sales_name:
            self.current_sales_id = self.db.insert_sales(sales_name, self.user_id)
            self.current_sales_name = sales_name
            self.update_sales_info()
            self.load_data()
            QMessageBox.information(self, "Sukses", f"Sales '{sales_name}' berhasil dibuat dan dipilih.")
            menu.close()

    def select_sales(self):
        sales_list = self.db.get_sales_list(self.user_id)
        if not sales_list:
            QMessageBox.warning(self, "Tidak Ada Sales", "Tidak ada sales yang tersedia. Silakan buat sales baru terlebih dahulu.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Pilih Sales")
        layout = QVBoxLayout(dialog)
        dialog.setMinimumWidth(400)
        dialog.setMinimumHeight(300)

        # Tambahkan input pencarian
        search_input = QLineEdit(dialog)
        search_input.setPlaceholderText("Cari sales...")
        layout.addWidget(search_input)

        # Gunakan QTableWidget dengan hanya satu kolom
        sales_table = QTableWidget(dialog)
        sales_table.setColumnCount(1)
        sales_table.setHorizontalHeaderLabels(["Nama Sales"])
        sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        sales_table.verticalHeader().setVisible(False)
        sales_table.setSelectionBehavior(QTableWidget.SelectRows)
        sales_table.setSelectionMode(QTableWidget.SingleSelection)
        layout.addWidget(sales_table)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        # Fungsi untuk mengisi tabel dengan data sales
        def populate_table(filter_text=""):
            sales_table.setRowCount(0)
            for sales in sales_list:
                if filter_text.lower() in sales[1].lower():
                    row = sales_table.rowCount()
                    sales_table.insertRow(row)
                    sales_table.setItem(row, 0, QTableWidgetItem(sales[1]))
                
                    # Set warna latar belakang selang-seling
                    if row % 2 == 0:
                        sales_table.item(row, 0).setBackground(QColor("#34485c"))

        # Hubungkan input pencarian ke fungsi filter
        search_input.textChanged.connect(populate_table)

        # Isi tabel awal
        populate_table()

        if dialog.exec_() == QDialog.Accepted and sales_table.currentRow() >= 0:
            selected_sales_name = sales_table.item(sales_table.currentRow(), 0).text()
            selected_sales = next(sales for sales in sales_list if sales[1] == selected_sales_name)
            self.current_sales_id = selected_sales[0]
            self.current_sales_name = selected_sales[1]
            self.update_sales_info()
            self.load_data()
            self.setting_button.show()

    def edit_sales(self, menu):
        if not self.current_sales_id:
            QMessageBox.warning(self, "No Sales Selected", "Please select a sales first.")
            return

        sales = self.db.get_sales(self.current_sales_id, self.user_id)
        
        sales_name, ok = QInputDialog.getText(self, "Edit Sales", "Nama Sales:", text=sales[1])
        if ok:
            self.db.update_sales(self.current_sales_id, sales_name, self.user_id)
            self.update_sales_info()
            QMessageBox.information(self, "Success", "Sales updated successfully.")
        
        menu.close()

    def delete_sales(self, menu):
        if not self.current_sales_id:
            QMessageBox.warning(self, "No Sales Selected", "Please select a sales first.")
            return

        warning_message = f'Anda akan menghapus sales "{self.current_sales_name}" dan semua proyek terkait.\n\n'
        warning_message += f'Ketik "HAPUS {self.current_sales_name}" untuk mengkonfirmasi penghapusan:'
    
        confirmation, ok = QInputDialog.getText(self, 'Konfirmasi Penghapusan', warning_message)
    
        if ok and confirmation == f"HAPUS {self.current_sales_name}":
            self.db.delete_sales(self.current_sales_id, self.user_id)
            self.current_sales_id = None
            self.current_sales_name = ""
            self.update_sales_info()
            self.load_data()
            self.select_sales_button.hide()
            QMessageBox.information(self, "Success", "Sales deleted successfully.")
        else:
            QMessageBox.information(self, "Cancelled", "Penghapusan sales dibatalkan.")
        
        menu.close()
    
    def load_data(self):
        self.table.setRowCount(0)  # Clear existing data
        if self.current_sales_id is not None:
            if self.is_viewing_history:
                self.select_sales_button.hide()
                self.setting_button.hide()
                projects = self.db.load_closed_book(self.current_book_name)
                for project in projects:
                    self.add_row(project[1:])  # Skip the ID column
            else:
                self.select_sales_button.show()
                self.setting_button.show()
                projects = self.db.get_sales_projects(self.current_sales_id, self.user_id)
                for project in projects:
                    self.add_row(project)
        self.update_total_commission()

    def add_row(self, data):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        for column, item in enumerate(data):
            if column in [self.total_project_column_index, self.commission_column_index, self.kb_column_index]:
                formatted_value = self.format_currency(item)
                self.table.setItem(row_position, column, QTableWidgetItem(formatted_value))
            else:
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))
    
        self.update_total_commission()


    def update_sales_info(self):
        if self.current_sales_id is not None:
            self.title_label.setText(f"Daftar Proyek Sales - {self.current_sales_name}")
            self.setting_button.show()  # Tampilkan tombol setting
        else:
            self.title_label.setText("Daftar Proyek Sales")
            self.setting_button.hide()  # Sembunyikan tombol setting
    
    def update_total_commission(self):
        total_commission = 0
        total_kb = 0
        for row in range(self.table.rowCount()):
            commission_item = self.table.item(row, self.commission_column_index)
            kb_item = self.table.item(row, self.kb_column_index)
            if commission_item and commission_item.text() and kb_item and kb_item.text():
                try:
                    commission_value = self.parse_currency(commission_item.text())
                    kb_value = self.parse_currency(kb_item.text())
                    total_commission += commission_value
                    total_kb += kb_value
                except ValueError:
                    print(f"Invalid commission or KB value at row {row}")
    
        net_commission = total_commission - total_kb
        self.total_commission_label.setText(f"Total Komisi: {self.format_currency(net_commission)}")
        self.total_kb_label.setText(f"Total KB: {self.format_currency(total_kb)}")
        self.total_commission = net_commission

    def open_add_dialog(self):
        if self.current_sales_id:
            dialog = AddSalesProjectDialog(self)
            if dialog.exec_():
                data = dialog.get_data()
                photo_path = dialog.get_photo_path()
                if not self.validate_numeric_input(data[3]) or not self.validate_numeric_input(data[4]) or not self.validate_numeric_input(data[5]):
                    QMessageBox.warning(self, "Invalid Input", "Total Proyek, Komisi, dan KB harus berupa angka.")
                    return
                if self.is_viewing_history:
                    new_id = self.db.add_to_closed_book(self.current_book_name, dict(zip(COLUMN_MAPPINGS[self.table_name].keys(), data)), photo_path, self.user_id, self.current_sales_id)
                    self.load_data()
                else:
                    now = datetime.now()
                    year, month = now.year, now.month
                    new_id = self.db.insert_sales_project(self.current_sales_id, data, year, month, self.user_id, photo_path)
                    formatted_data = list(data)
                    formatted_data[3] = self.format_currency(data[3])  # Format Total Proyek
                    formatted_data[4] = self.format_currency(data[4])  # Format Komisi
                    formatted_data[5] = self.format_currency(data[5])  # Format KB
                    self.add_row((new_id,) + tuple(formatted_data))
                    self.load_data()  # Reload data to show the new entry
                self.update_total_commission()
                QMessageBox.information(self, "Sukses", "Data proyek sales berhasil ditambahkan.")
        else:
            QMessageBox.warning(self, "No Sales Selected", "Please select or create a sales first.")

    def open_edit_dialog(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin diedit terlebih dahulu.")
            return

        selected_row = selected_items[0].row()
        dialog = AddSalesProjectDialog(self)

        initial_data = []
        for col in range(1, self.table.columnCount()):  # Start from 1 to skip ID column
            item = self.table.item(selected_row, col)
            if item is not None:
                if col in [4, 5, 6]:  # Total Proyek, Komisi, dan KB columns
                    initial_data.append(self.parse_currency(item.text()))
                else:
                    initial_data.append(item.text())
            else:
                initial_data.append("")

        dialog.load_data(initial_data)

        if dialog.exec_():
            if dialog.validate_data():
                data = dialog.get_data()
                photo_path = dialog.get_photo_path()
                for col, item_text in enumerate(data, start=1):  # Start from 1 to skip ID column
                    if col in [4, 5, 6]:  # Total Proyek, Komisi, dan KB columns
                        parsed_value = self.parse_currency(item_text)
                        formatted_value = self.format_currency(parsed_value)
                        self.table.setItem(selected_row, col, QTableWidgetItem(formatted_value))
                    else:
                        self.table.setItem(selected_row, col, QTableWidgetItem(str(item_text) if item_text is not None else ""))
    
                if self.is_viewing_history:
                    closed_book_data = self.db.load_closed_book(self.current_book_name)
                    if selected_row < len(closed_book_data):
                        record_id = closed_book_data[selected_row][0]
                        self.db.update_in_closed_book(self.current_book_name, record_id, dict(zip(COLUMN_MAPPINGS[self.table_name].keys(), data)), photo_path, self.user_id, self.current_sales_id)
                    else:
                        QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan diupdate.")
                        return
                else:
                    sales_projects = self.db.get_sales_projects(self.current_sales_id, self.user_id)
                    if selected_row < len(sales_projects):
                        project_id = sales_projects[selected_row][0]
                        self.db.update_sales_project(project_id, data, self.user_id, photo_path)
                    else:
                        QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan diupdate.")
                        return
        
                self.update_total_commission()
                QMessageBox.information(self, "Sukses", "Data proyek sales berhasil diedit.")
            else:
                QMessageBox.warning(self, "Input Tidak Valid", "Total Proyek, Komisi, dan KB harus berupa angka.")

        self.load_data()  # Reload data to reflect changes

    def validate_numeric_input(self, value):
        try:
            float(str(value).replace('Rp', '').replace('.', '').replace(',', '').strip())
            return True
        except ValueError:
            return False

    def format_currency(self, value):
        print(f"Formatting currency for value: {value}")  # Debugging statement
        try:
            # Remove any existing formatting and convert to float
            if isinstance(value, str):
                value = self.parse_currency(value)
            # Round to nearest integer and format without decimal places
            rounded_value = int(value)  # Convert to int to avoid decimal-related issues
            return f"Rp {rounded_value:,}".replace(',', '.')
        except ValueError:
            # If conversion fails, return the original value
            return str(value)

    def parse_currency(self, value):
        print(f"Parsing currency for value: {value}")  # Debugging statement
        # Remove currency symbol and separators, then convert to float
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        try:
            return float(cleaned_value) if '.' in cleaned_value else int(cleaned_value)
        except ValueError:
            return 0.0


    def delete_selected_row(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin dihapus terlebih dahulu.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus data ini?',
                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            selected_row = selected_items[0].row()

            if self.is_viewing_history:
                closed_book_data = self.db.load_closed_book(self.current_book_name)
                if selected_row < len(closed_book_data):
                    record_id = closed_book_data[selected_row][0]
                    photo_path = closed_book_data[selected_row][-1]  # Assuming photo_path is the last column
                
                    # Delete photo if it exists
                    if photo_path and os.path.exists(photo_path):
                        os.remove(photo_path)
                    
                        # Delete the parent folder if it's empty
                        parent_folder = os.path.dirname(photo_path)
                        if not os.listdir(parent_folder):
                            os.rmdir(parent_folder)
                
                    self.db.delete_from_closed_book(self.current_book_name, record_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                    return
            else:
                sales_projects = self.db.get_sales_projects(self.current_sales_id, self.user_id)
                if selected_row < len(sales_projects):
                    record_id = sales_projects[selected_row][0]
                    photo_path = self.db.get_sales_project_photo(record_id, self.user_id)
                
                    # Delete photo if it exists
                    if photo_path and os.path.exists(photo_path):
                        os.remove(photo_path)
                    
                        # Delete the parent folder if it's empty
                        parent_folder = os.path.dirname(photo_path)
                        if not os.listdir(parent_folder):
                            os.rmdir(parent_folder)
                
                    self.db.delete_record(self.table_name, record_id, self.user_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                    return

            self.table.removeRow(selected_row)
            QMessageBox.information(self, "Hapus Data", "Data berhasil dihapus.")
            self.update_total_commission()

    def close_book(self):
        if self.current_sales_id is None:
            QMessageBox.warning(self, 'Tutup Buku', 'Silakan pilih sales terlebih dahulu.')
            return

        if self.table.rowCount() == 0:
            QMessageBox.warning(self, 'Tutup Buku', 'Tidak dapat menutup buku karena data kosong.')
            return

        reply = QMessageBox.question(self, 'Tutup Buku', f'Anda yakin ingin menutup buku untuk sales {self.current_sales_name}? Ini akan membuat backup data saat ini dan menghapus semua data dari tabel.',
                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            backup_name = self.db.close_book_for_person(self.table_name, self.current_sales_id, self.user_id)
            self.load_data()  # Reload the (now empty) table
            QMessageBox.information(self, "Tutup Buku", f"Buku untuk sales {self.current_sales_name} telah ditutup. Data lama telah dibackup ke {backup_name} dan tabel telah direset.")

    def view_history(self):
        if self.current_sales_id is None:
            QMessageBox.warning(self, 'Lihat Riwayat', 'Silakan pilih sales terlebih dahulu.')
            return

        closed_books = self.db.get_closed_books_for_person(self.table_name, self.current_sales_id, self.user_id)
        if closed_books:
            formatted_names = {self.format_backup_name(book): book for book in closed_books}
        
            formatted_name, ok = QInputDialog.getItem(
                self, 
                "Pilih Riwayat", 
                f"Riwayat {self.current_sales_name}:", 
                sorted(formatted_names.keys(), reverse=True),
                0, 
                False
            )
        
            if ok:
                original_name = formatted_names[formatted_name]
                data = self.db.load_closed_book(original_name)
                self.display_history(data, formatted_name, original_name)
                self.load_data()
        else:
            QMessageBox.information(self, "Tidak Ada Data", f"Tidak ada riwayat tersimpan untuk sales {self.current_sales_name}.")

    def format_backup_name(self, backup_name):
        parts = backup_name.split('_')
        year = parts[-4]
        month = parts[-3]
        day = parts[-2]
        count = parts[-1]
        
        month_name = datetime(int(year), int(month), 1).strftime('%B')
        
        return f"{self.current_sales_name} {year} {month_name} {day} ({count})"
    
    def display_history(self, data, book_name, original_name):
        self.table.setRowCount(0)  # Clear the current table
        self.title_label.setText(f"Riwayat {self.title_label.text().split(' - ')[0]} - {book_name}")

        for row_data in data:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            for column, item in enumerate(row_data[1:]):  # Exclude id
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))

        self.close_book_button.hide()
        self.view_history_button.hide()
        self.setting_button.hide()

        self.return_button = QPushButton("Kembali ke Data Saat Ini")
        self.return_button.clicked.connect(self.return_to_current_data)
        self.button_layout.addWidget(self.return_button)

        self.current_book_name = original_name
        self.is_viewing_history = True
        
        self.update_total_commission()


    def return_to_current_data(self):
        self.title_label.setText(f"Daftar Proyek Sales - {self.current_sales_name}")
        self.is_viewing_history = False
        self.current_book_name = None
        self.load_data()
        self.close_book_button.show()
        self.view_history_button.show()
        if self.current_sales_id:
           self.setting_button.show()  # Tampilkan kembali tombol setting jika ada sales yang dipilih
        else:
           self.setting_button.hide()
        self.add_button.setEnabled(True)
        self.edit_button.setEnabled(True)
        self.delete_button.setEnabled(True)
        self.return_button.setParent(None)
        self.return_button = None
    
    def calculate_totals(self):
        total_commission = 0
        total_kb = 0
        for row in range(self.table.rowCount()):
            commission_item = self.table.item(row, self.commission_column_index)
            kb_item = self.table.item(row, self.kb_column_index)
            if commission_item and commission_item.text() and kb_item and kb_item.text():
                total_commission += self.parse_currency(commission_item.text())
                total_kb += self.parse_currency(kb_item.text())
        return total_commission, total_kb

    def export_to_excel(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if file_name:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Write headers
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount()) if not self.table.isColumnHidden(i)]
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Write data
            for row in range(self.table.rowCount()):
                col_index = 1
                for column in range(self.table.columnCount()):
                    if not self.table.isColumnHidden(column):
                        item = self.table.item(row, column)
                        if item is not None:
                            cell = sheet.cell(row=row+2, column=col_index, value=item.text())
                            if headers[col_index-1] == "Keterangan":
                                cell.alignment = Alignment(wrapText=True, vertical='top')
                                cell = self.format_keterangan_cell(cell)
                        else:
                            sheet.cell(row=row+2, column=col_index, value='')
                        col_index += 1
            
            sheet.append([""] * sheet.max_column)
            
            total_commission, total_kb = self.calculate_totals()
            row = sheet.max_row + 2
            sheet.cell(row=row, column=1, value="Total Komisi").font = Font(bold=True)
            sheet.cell(row=row, column=2, value=self.format_currency(total_commission - total_kb))
            row += 1
            sheet.cell(row=row, column=1, value="Total KB").font = Font(bold=True)
            sheet.cell(row=row, column=2, value=self.format_currency(total_kb))
            
            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_name)
            QMessageBox.information(self, "Export Successful", f"Data has been exported to {file_name}")

class TukangTable(TableWidget):
    def __init__(self, parent=None):
        super().__init__("worker_projects", parent)
        setup_error_handling()
        self.title_label.setText("Daftar Proyek Tukang")
        self.setup_table()
        self.setup_setting_button()
        self.setup_view_photo_button()
        self.current_tukang_id = None
        self.current_tukang_name = ""
        self.size_column_index = 4
        self.kb_column_index = 5

    def setup_table(self):
        self.table.setColumnCount(len(COLUMN_MAPPINGS[self.table_name]) + 1)  # +1 for ID column
        self.table.setHorizontalHeaderLabels(["ID"] + list(COLUMN_MAPPINGS[self.table_name].keys()))
        self.table.hideColumn(0)  # Hide ID column
    
    def setup_setting_button(self):
        self.select_tukang_button = QPushButton("Pilih Tukang")
        self.select_tukang_button.clicked.connect(self.select_tukang)
        self.button_layout.addWidget(self.select_tukang_button)

        self.setting_button = QPushButton("Setting")
        self.setting_button.clicked.connect(self.open_settings)
        self.button_layout.addWidget(self.setting_button)
        self.setting_button.hide()  # Hide the button initially
    
    def open_settings(self):
        menu = QDialog(self)
        menu.setWindowTitle("Setting Tukang")
        menu.setMinimumWidth(400)
        layout = QVBoxLayout(menu)

        new_tukang_button = QPushButton("Tukang Baru")
        new_tukang_button.clicked.connect(lambda: self.create_new_tukang(menu))
        layout.addWidget(new_tukang_button)

        edit_tukang_button = QPushButton("Edit Tukang")
        edit_tukang_button.clicked.connect(lambda: self.edit_tukang(menu))
        layout.addWidget(edit_tukang_button)

        delete_tukang_button = QPushButton("Hapus Tukang")
        delete_tukang_button.clicked.connect(lambda: self.delete_tukang(menu))
        layout.addWidget(delete_tukang_button)

        # Disable edit and delete buttons if no tukang is selected
        edit_tukang_button.setEnabled(self.current_tukang_id is not None)
        delete_tukang_button.setEnabled(self.current_tukang_id is not None)

        menu.setLayout(layout)
        menu.exec_()

    
    def add_row(self, data):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        for column, item in enumerate(data):
            if column == self.kb_column_index:  # KB column
                formatted_value = self.format_currency(item)
                self.table.setItem(row_position, column, QTableWidgetItem(formatted_value))
            else:
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))

    def load_data(self):
        self.table.setRowCount(0)  # Clear existing data
        if self.current_tukang_id:
            if self.is_viewing_history:
                projects = self.db.load_closed_book(self.current_book_name)
                for project in projects:
                    self.add_row(project[1:])  # Skip the ID column
            else:
                projects = self.db.get_worker_projects(self.current_tukang_id, self.user_id)
                for project in projects:
                    self.add_row(project)

    def create_new_tukang(self):
        tukang_name, ok = QInputDialog.getText(self, "Tukang Baru", "Nama Tukang:")
        if ok and tukang_name:
            self.current_tukang_id = self.db.insert_tukang(tukang_name, self.user_id)
            self.update_tukang_info()
            self.load_data()
            QMessageBox.information(self, "Sukses", f"Tukang '{tukang_name}' berhasil dibuat dan dipilih.")

    def select_tukang(self):
        tukang_list = self.db.get_tukang_list(self.user_id)
        if not tukang_list:
            QMessageBox.warning(self, "Tidak Ada Tukang", "Tidak ada tukang yang tersedia. Silakan buat tukang baru terlebih dahulu.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Pilih Tukang")
        layout = QVBoxLayout(dialog)
        dialog.setMinimumWidth(400)
        dialog.setMinimumHeight(300)

        # Tambahkan input pencarian
        search_input = QLineEdit(dialog)
        search_input.setPlaceholderText("Cari tukang...")
        layout.addWidget(search_input)

        # Gunakan QTableWidget dengan hanya satu kolom
        tukang_table = QTableWidget(dialog)
        tukang_table.setColumnCount(1)
        tukang_table.setHorizontalHeaderLabels(["Nama Tukang"])
        tukang_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tukang_table.verticalHeader().setVisible(False)
        tukang_table.setSelectionBehavior(QTableWidget.SelectRows)
        tukang_table.setSelectionMode(QTableWidget.SingleSelection)
        layout.addWidget(tukang_table)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        # Fungsi untuk mengisi tabel dengan data tukang
        def populate_table(filter_text=""):
            tukang_table.setRowCount(0)
            for tukang in tukang_list:
                if filter_text.lower() in tukang[1].lower():
                    row = tukang_table.rowCount()
                    tukang_table.insertRow(row)
                    tukang_table.setItem(row, 0, QTableWidgetItem(tukang[1]))
                
                    # Set warna latar belakang selang-seling
                    if row % 2 == 0:
                        tukang_table.item(row, 0).setBackground(QColor("#34485c"))

        # Hubungkan input pencarian ke fungsi filter
        search_input.textChanged.connect(populate_table)

        # Isi tabel awal
        populate_table()

        if dialog.exec_() == QDialog.Accepted and tukang_table.currentRow() >= 0:
            selected_tukang_name = tukang_table.item(tukang_table.currentRow(), 0).text()
            selected_tukang = next(tukang for tukang in tukang_list if tukang[1] == selected_tukang_name)
            self.current_tukang_id = selected_tukang[0]
            self.current_tukang_name = selected_tukang[1]
            self.update_tukang_info()
            self.load_data()
            self.setting_button.show()

    def edit_tukang(self):
        if not self.current_tukang_id:
            QMessageBox.warning(self, "No Tukang Selected", "Please select a tukang first.")
            return

        tukang = self.db.get_tukang(self.current_tukang_id, self.user_id)
        
        tukang_name, ok = QInputDialog.getText(self, "Edit Tukang", "Nama Tukang:", text=tukang[1])
        if ok:
            self.db.update_tukang(self.current_tukang_id, tukang_name, self.user_id)
            self.update_tukang_info()
            QMessageBox.information(self, "Success", "Tukang updated successfully.")

    def delete_tukang(self, menu=None):
        if not self.current_tukang_id:
            QMessageBox.warning(self, "No Tukang Selected", "Please select a tukang first.")
            return

        warning_message = f'Anda akan menghapus tukang "{self.current_tukang_name}" dan semua proyek terkait.\n\n'
        warning_message += f'Ketik "HAPUS {self.current_tukang_name}" untuk mengkonfirmasi penghapusan:'
    
        confirmation, ok = QInputDialog.getText(self, 'Konfirmasi Penghapusan', warning_message)
    
        if ok and confirmation == f"HAPUS {self.current_tukang_name}":
            self.db.delete_tukang(self.current_tukang_id, self.user_id)
            self.current_tukang_id = None
            self.current_tukang_name = ""
            self.update_tukang_info()
            self.load_data()
            self.setting_button.hide()  # Add this line
            QMessageBox.information(self, "Success", "Tukang deleted successfully.")
        else:
            QMessageBox.information(self, "Cancelled", "Penghapusan tukang dibatalkan.")

        if menu:
            menu.close()

    def update_tukang_info(self):
        if self.current_tukang_id:
            tukang = self.db.get_tukang(self.current_tukang_id, self.user_id)
            self.current_tukang_name = tukang[1]
            self.title_label.setText(f"Daftar Proyek Tukang - {self.current_tukang_name}")
            self.setting_button.show()  # Add this line
        else:
            self.title_label.setText("Daftar Proyek Tukang")
            self.setting_button.hide()  # Add this line
    
    def setup_view_photo_button(self):
        self.view_photo_button = QPushButton("Lihat Gambar")
        self.view_photo_button.clicked.connect(self.view_photo)
        self.view_photo_button.hide()
        self.button_layout.addWidget(self.view_photo_button)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)

    def on_selection_changed(self):
        selected_items = self.table.selectedItems()
        if selected_items:
            self.view_photo_button.show()
        else:
            self.view_photo_button.hide()

    def view_photo(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()

        if self.is_viewing_history:
            closed_book_data = self.db.load_closed_book(self.current_book_name)
            if selected_row < len(closed_book_data):
                project_id = closed_book_data[selected_row][0]
                photo_path = closed_book_data[selected_row][-1]  # Assuming photo_path is the last column
            else:
                QMessageBox.warning(self, "Error", "Tidak dapat menemukan data proyek.")
                return
        else:
            project_id = self.table.item(selected_row, 0).text()  # Assuming ID is in the first column
            photo_path = self.db.get_worker_project_photo(project_id, self.user_id)

        if photo_path and os.path.exists(photo_path):
            dialog = PhotoViewerDialog(self, photo_path)
            dialog.exec_()
        else:
            QMessageBox.information(self, "Tidak Ada Gambar", "Tidak ada gambar tersedia untuk proyek ini.")

    def open_add_dialog(self):
        if self.current_tukang_id:
            dialog = AddTukangProjectDialog(self)
            if dialog.exec_():
                data = dialog.get_data()
                photo_path = dialog.get_photo_path()
                if not self.validate_numeric_input(data[4]):  # KB validation
                    QMessageBox.warning(self, "Invalid Input", "KB harus berupa angka.")
                    return
                if self.is_viewing_history:
                    new_id = self.db.add_to_closed_book(self.current_book_name, dict(zip(COLUMN_MAPPINGS[self.table_name].keys(), data)), photo_path, self.user_id, self.current_tukang_id)
                    self.load_data()
                else:
                    now = datetime.now()
                    year, month = now.year, now.month
                    new_id = self.db.insert_worker_project(self.current_tukang_id, data, year, month, self.user_id, photo_path)
                    formatted_data = list(data)
                    formatted_data[4] = self.format_currency(data[4])  # Format KB
                    self.add_row((new_id,) + tuple(formatted_data))
                QMessageBox.information(self, "Sukses", "Data proyek tukang berhasil ditambahkan.")
        else:
            QMessageBox.warning(self, "No Tukang Selected", "Please select or create a tukang first.")

    def open_edit_dialog(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin diedit terlebih dahulu.")
            return

        selected_row = selected_items[0].row()
        dialog = AddTukangProjectDialog(self)

        initial_data = []
        for col in range(1, self.table.columnCount()):  # Start from 1 to skip ID column
            item = self.table.item(selected_row, col)
            if item is not None:
                if col == self.kb_column_index:
                    initial_data.append(str(self.parse_currency(item.text())))
                else:
                    initial_data.append(item.text())
            else:
                initial_data.append("")

        dialog.load_data(initial_data)

        if dialog.exec_():
            data = dialog.get_data()
            photo_path = dialog.get_photo_path()
            if not self.validate_numeric_input(data[4]):  # KB validation
                QMessageBox.warning(self, "Invalid Input", "KB harus berupa angka.")
                return
        
            for col, item_text in enumerate(data, start=1):  # Start from 1 to skip ID column
                if col == self.kb_column_index:
                    parsed_value = self.parse_currency(item_text)
                    formatted_value = self.format_currency(parsed_value)
                    self.table.setItem(selected_row, col, QTableWidgetItem(formatted_value))
                else:
                    self.table.setItem(selected_row, col, QTableWidgetItem(str(item_text)))

            if self.is_viewing_history:
                closed_book_data = self.db.load_closed_book(self.current_book_name)
                if selected_row < len(closed_book_data):
                    record_id = closed_book_data[selected_row][0]
                    self.db.update_in_closed_book(self.current_book_name, record_id, dict(zip(COLUMN_MAPPINGS[self.table_name].keys(), data)), photo_path, self.user_id, self.current_tukang_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan diupdate.")
                    return
            else:
                project_id = int(self.table.item(selected_row, 0).text())  # Get ID from hidden column
                self.db.update_worker_project(project_id, data, self.user_id, photo_path)
        
            self.load_data()  # Reload data to reflect changes
            QMessageBox.information(self, "Sukses", "Data proyek tukang berhasil diedit.")

    def validate_numeric_input(self, value):
        # Hapus format mata uang dan separator ribuan
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        # Periksa apakah string hanya terdiri dari digit
        return cleaned_value.isdigit()

    def format_currency(self, value):
        try:
            # Remove any existing formatting and convert to float
            if isinstance(value, str):
                value = self.parse_currency(value)
            # Round to nearest integer and format without decimal places
            rounded_value = int(value)  # Convert to int to avoid decimal-related issues
            return f"Rp {rounded_value:,}".replace(',', '.')
        except ValueError:
            # If conversion fails, return the original value
            return str(value)

    def parse_currency(self, value):
        # Remove currency symbol and separators, then convert to float
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        try:
            return int(float(cleaned_value))
        except ValueError:
            return 0

    def delete_selected_row(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin dihapus terlebih dahulu.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus data ini?',
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            selected_row = selected_items[0].row()

            if self.is_viewing_history:
                closed_book_data = self.db.load_closed_book(self.current_book_name)
                if selected_row < len(closed_book_data):
                    record_id = closed_book_data[selected_row][0]
                    photo_path = closed_book_data[selected_row][-1]  # Assuming photo_path is the last column
                
                    # Delete photo if it exists
                    if photo_path and os.path.exists(photo_path):
                        os.remove(photo_path)
                    
                        # Delete the parent folder if it's empty
                        parent_folder = os.path.dirname(photo_path)
                        if not os.listdir(parent_folder):
                            os.rmdir(parent_folder)

                    self.db.delete_from_closed_book(self.current_book_name, record_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                    return
            else:
                if self.current_tukang_id is None:
                    QMessageBox.warning(self, "Error", "Tidak ada tukang yang dipilih.")
                    return
                projects = self.db.get_worker_projects(self.current_tukang_id, self.user_id)
                if selected_row < len(projects):
                    record_id = projects[selected_row][0]
                    photo_path = self.db.get_worker_project_photo(record_id, self.user_id)
                
                    # Delete photo if it exists
                    if photo_path and os.path.exists(photo_path):
                        os.remove(photo_path)
                    
                        # Delete the parent folder if it's empty
                        parent_folder = os.path.dirname(photo_path)
                        if not os.listdir(parent_folder):
                            os.rmdir(parent_folder)
                            
                    self.db.delete_record(self.table_name, record_id, self.user_id)
                else:
                    QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                    return

            self.table.removeRow(selected_row)
            QMessageBox.information(self, "Hapus Data", "Data berhasil dihapus.")

    def close_book(self):
        if self.current_tukang_id is None:
            QMessageBox.warning(self, 'Tutup Buku', 'Silakan pilih tukang terlebih dahulu.')
            return

        if self.table.rowCount() == 0:
            QMessageBox.warning(self, 'Tutup Buku', 'Tidak dapat menutup buku karena data kosong.')
            return

        reply = QMessageBox.question(self, 'Tutup Buku', f'Anda yakin ingin menutup buku untuk tukang {self.current_tukang_name}? Ini akan membuat backup data saat ini dan menghapus semua data dari tabel.',
                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            backup_name = self.db.close_book_for_person(self.table_name, self.current_tukang_id, self.user_id)
            self.load_data()  # Reload the (now empty) table
            QMessageBox.information(self, "Tutup Buku", f"Buku untuk tukang {self.current_tukang_name} telah ditutup. Data lama telah dibackup ke {backup_name} dan tabel telah direset.")

    def view_history(self):
        if self.current_tukang_id is None:
            QMessageBox.warning(self, 'Lihat Riwayat', 'Silakan pilih tukang terlebih dahulu.')
            return

        closed_books = self.db.get_closed_books_for_person(self.table_name, self.current_tukang_id, self.user_id)
        if closed_books:
            formatted_names = {self.format_backup_name(book): book for book in closed_books}
        
            formatted_name, ok = QInputDialog.getItem(
                self, 
                "Pilih Riwayat", 
                f"Riwayat {self.current_tukang_name}:", 
                sorted(formatted_names.keys(), reverse=True),
                0, 
                False
            )
        
            if ok:
                original_name = formatted_names[formatted_name]
                data = self.db.load_closed_book(original_name)
                self.display_history(data, formatted_name, original_name)
                self.load_data()
        else:
            QMessageBox.information(self, "Tidak Ada Data", f"Tidak ada riwayat tersimpan untuk tukang {self.current_tukang_name}.")

    def format_backup_name(self, backup_name):
        parts = backup_name.split('_')
        year = parts[-4]
        month = parts[-3]
        day = parts[-2]
        count = parts[-1]
        
        month_name = datetime(int(year), int(month), 1).strftime('%B')
        
        return f"{self.current_tukang_name} {year} {month_name} {day} ({count})"

    def display_history(self, data, book_name, original_name):
        self.table.setRowCount(0)  # Clear the current table
        self.title_label.setText(f"Riwayat {self.title_label.text().split(' - ')[0]} - {book_name}")
    
        for row_data in data:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            for column, item in enumerate(row_data[1:]):  # Exclude id
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))
    
        self.close_book_button.hide()
        self.view_history_button.hide()
        self.setting_button.hide()
        self.select_tukang_button.hide()

        self.return_button = QPushButton("Kembali ke Data Saat Ini")
        self.return_button.clicked.connect(self.return_to_current_data)
        self.button_layout.addWidget(self.return_button)

        self.current_book_name = original_name
        self.is_viewing_history = True

    def return_to_current_data(self):
        self.title_label.setText(f"Daftar Proyek Tukang - {self.current_tukang_name}")
        self.is_viewing_history = False
        self.current_book_name = None
        self.load_data()
        self.close_book_button.show()
        self.view_history_button.show()
        self.select_tukang_button.show()
        self.add_button.setEnabled(True)
        self.edit_button.setEnabled(True)
        self.delete_button.setEnabled(True)
        self.return_button.setParent(None)
        self.return_button = None

class MaterialTable(TableWidget):
    def __init__(self, parent=None):
        super().__init__("materials_usage", parent)
        self.title_label.setText("Daftar Pemakaian Bahan")
        self.setup_project_info_display()
        self.setup_table()
        self.setup_project_buttons()
        self.setup_total_labels()
        self.current_project_id = None
        self.current_project_name = ""
        self.total_project = 0

    def setup_project_info_display(self):
        self.project_info_widget = QWidget()
        self.project_info_layout = QVBoxLayout(self.project_info_widget)
        
        # Centered project name label
        self.project_name_label = QLabel("Belum ada proyek dipilih")
        self.project_name_label.setFont(QFont("Arial", 14, QFont.Bold))
        self.project_name_label.setAlignment(Qt.AlignCenter)
        self.project_info_layout.addWidget(self.project_name_label)
        
        # Horizontal line
        self.line = QFrame()
        self.line.setFrameShape(QFrame.HLine)
        self.line.setFrameShadow(QFrame.Sunken)
        self.project_info_layout.addWidget(self.line)

        # Add vertical spacer for padding
        spacer = QSpacerItem(20, 10, QSizePolicy.Minimum, QSizePolicy.Fixed)
        self.project_info_layout.addItem(spacer)

        self.project_details_layout = QHBoxLayout()
        self.project_info_layout.addLayout(self.project_details_layout)

        self.sales_label = QLabel()
        self.worker_label = QLabel()
        self.date_label = QLabel()
        self.total_label = QLabel()
        self.dp_label = QLabel()

        self.project_details_layout.addWidget(self.sales_label)
        self.project_details_layout.addWidget(self.worker_label)
        self.project_details_layout.addWidget(self.date_label)
        self.project_details_layout.addWidget(self.total_label)
        self.project_details_layout.addWidget(self.dp_label)

        # Insert the project info widget after the search layout
        self.layout.insertWidget(2, self.project_info_widget)

    def setup_table(self):
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Tanggal", "Nama Barang", "Quantity", "Harga Satuan", "Total", "Keterangan", "ID"])
        self.table.hideColumn(6)  # Hide the ID column

    def setup_project_buttons(self):
        self.new_project_button = QPushButton("Proyek Baru")
        self.new_project_button.clicked.connect(self.create_new_project)
        self.button_layout.addWidget(self.new_project_button)

        self.select_project_button = QPushButton("Pilih Proyek")
        self.select_project_button.clicked.connect(self.select_project)
        self.button_layout.addWidget(self.select_project_button)

        self.edit_project_button = QPushButton("Edit Proyek")
        self.edit_project_button.clicked.connect(self.edit_project)
        self.button_layout.addWidget(self.edit_project_button)
        self.edit_project_button.hide()

        self.delete_project_button = QPushButton("Hapus Proyek")
        self.delete_project_button.clicked.connect(self.delete_project)
        self.button_layout.addWidget(self.delete_project_button)
        self.delete_project_button.hide()

    def setup_total_labels(self):
        self.total_price_label = QLabel("Total Harga: Rp 0")
        self.total_price_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        self.layout.addWidget(self.total_price_label)

        self.total_profit_label = QLabel("Total Keuntungan: Rp 0")
        self.total_profit_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        self.layout.addWidget(self.total_profit_label)

    def load_data(self):
        self.table.setRowCount(0)
        if self.current_project_id:
            materials = self.db.get_material_usage(self.current_project_id, self.user_id)
            for material in materials:
                data = list(material[2:])  # Exclude id and project_id
                data[3] = self.format_price(data[3])  # Format Harga Satuan
                self.add_row(data)
        self.update_total_price()
    
    def get_project_input(self, title, initial_data=None):
        project_data = {}
        fields = [
            ("Nama Proyek", "text"),
            ("Nama Sales", "text"),
            ("Nama Tukang", "text"),
            ("Tanggal Mulai", "date"),
            ("Tanggal Selesai", "date"),
            ("Total Proyek", "text"),
            ("DP", "text")
        ]

        for field, input_type in fields:
            if input_type == "text":
                value, ok = QInputDialog.getText(self, title, field, text=initial_data.get(field, "") if initial_data else "")
            elif input_type == "date":
                date_dialog = DateInputDialog(self, field)
                if initial_data and field in initial_data:
                    date_dialog.date_edit.setDate(QDate.fromString(initial_data[field], "dd/MM/yyyy"))
                if date_dialog.exec_() == QDialog.Accepted:
                    value = date_dialog.get_date().toString("dd/MM/yyyy")
                    ok = True
                else:
                    ok = False
            
            if ok and value:
                project_data[field] = value
            else:
                return None  # User canceled or left a field empty

        return project_data

    def select_project(self):
        projects = self.db.get_projects(self.user_id)
        if not projects:
            QMessageBox.warning(self, "Tidak Ada Proyek", "Tidak ada proyek yang tersedia. Silakan buat proyek baru terlebih dahulu.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Pilih Proyek")
        layout = QVBoxLayout(dialog)
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)

        # Tambahkan input pencarian
        search_input = QLineEdit(dialog)
        search_input.setPlaceholderText("Cari proyek...")
        layout.addWidget(search_input)

        # Gunakan QTableWidget alih-alih QListWidget
        project_table = QTableWidget(dialog)
        project_table.setColumnCount(2)
        project_table.setHorizontalHeaderLabels(["Nama Proyek", "Tanggal Mulai"])
        project_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        project_table.verticalHeader().setVisible(False)
        project_table.setSelectionBehavior(QTableWidget.SelectRows)
        project_table.setSelectionMode(QTableWidget.SingleSelection)
        layout.addWidget(project_table)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        # Fungsi untuk mengisi tabel dengan data proyek
        def populate_table(filter_text=""):
            project_table.setRowCount(0)
            for i, project in enumerate(projects):
                if filter_text.lower() in project[1].lower():
                    row = project_table.rowCount()
                    project_table.insertRow(row)
                    project_table.setItem(row, 0, QTableWidgetItem(project[1]))
                    project_table.setItem(row, 1, QTableWidgetItem(project[4]))
                
                    # Set warna latar belakang selang-seling
                    if row % 2 == 0:
                        project_table.item(row, 0).setBackground(QColor("#34485c"))
                        project_table.item(row, 1).setBackground(QColor("#34485c"))

        # Hubungkan input pencarian ke fungsi filter
        search_input.textChanged.connect(populate_table)

        # Isi tabel awal
        populate_table()

        if dialog.exec_() == QDialog.Accepted and project_table.currentRow() >= 0:
            selected_project_name = project_table.item(project_table.currentRow(), 0).text()
            selected_project = next(project for project in projects if project[1] == selected_project_name)
            self.current_project_id = selected_project[0]
            self.update_project_info()
            self.load_data()
            self.edit_project_button.show()
            self.delete_project_button.show()

    def format_currency(self, value):
        try:
            if isinstance(value, str):
                value = self.parse_currency(value)
            rounded_value = int(value)
            return f"Rp {rounded_value:,}".replace(',', '.')
        except ValueError:
            return str(value)

    def parse_currency(self, value):
        cleaned_value = str(value).replace('Rp', '').replace('.', '').replace(',', '').strip()
        try:
            return int(float(cleaned_value))
        except ValueError:
            return 0

    def create_new_project(self):
        dialog = ProjectInputDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            project_data = dialog.get_project_data()
            
            # Cek apakah nama proyek sudah ada
            existing_projects = self.db.get_projects(self.user_id)
            existing_project_names = [project[1].lower() for project in existing_projects]
            
            if project_data['Nama Proyek'].lower() in existing_project_names:
                QMessageBox.warning(self, "Nama Proyek Sudah Ada", "Proyek dengan nama yang sama sudah ada. Silakan gunakan nama lain.")
                return
            
            self.current_project_id = self.db.insert_project(tuple(project_data.values()), self.user_id)
            self.update_project_info()
            self.load_data()
            self.edit_project_button.show()
            self.delete_project_button.show()
            QMessageBox.information(self, "Sukses", f"Proyek '{project_data['Nama Proyek']}' berhasil dibuat dan dipilih.")

    def edit_project(self):
        if not self.current_project_id:
            QMessageBox.warning(self, "No Project Selected", "Please select a project first.")
            return

        projects = self.db.get_projects(self.user_id)
        project = next((p for p in projects if p[0] == self.current_project_id), None)
        if not project:
            QMessageBox.warning(self, "Error", "Proyek tidak ditemukan.")
            return

        initial_data = {
            "Nama Proyek": project[1],
            "Nama Sales": project[2],
            "Nama Tukang": project[3],
            "Tanggal Mulai": project[4],
            "Tanggal Selesai": project[5],
            "Total Proyek": project[6],
            "DP": project[7]
        }

        dialog = ProjectInputDialog(self, initial_data)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_project_data()
        
            if updated_data['Nama Proyek'].lower() != project[1].lower():
                existing_project_names = [p[1].lower() for p in projects if p[0] != self.current_project_id]
                if updated_data['Nama Proyek'].lower() in existing_project_names:
                    QMessageBox.warning(self, "Nama Proyek Sudah Ada", "Proyek dengan nama yang sama sudah ada. Silakan gunakan nama lain.")
                    return
        
            updated_project_data = tuple(updated_data.values()) + (self.current_project_id,)
            self.db.update_project(updated_project_data, self.user_id)
            self.update_project_info()
            QMessageBox.information(self, "Success", "Project updated successfully.")

    def delete_project(self):
        if not self.current_project_id:
            QMessageBox.warning(self, "No Project Selected", "Please select a project first.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus proyek ini? Semua data pemakaian bahan terkait juga akan dihapus.',
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            confirmation, ok = QInputDialog.getText(
                self, 
                'Konfirmasi Penghapusan', 
                f'Ketik "HAPUS {self.current_project_name}" untuk mengkonfirmasi penghapusan:'
            )
        
            if ok and confirmation == f"HAPUS {self.current_project_name}":
                self.db.delete_project(self.current_project_id, self.user_id)
                self.current_project_id = None
                self.current_project_name = ""
                self.update_project_info()
                self.load_data()
                self.edit_project_button.hide()
                self.delete_project_button.hide()
                QMessageBox.information(self, "Success", "Project deleted successfully.")
            else:
                QMessageBox.information(self, "Cancelled", "Penghapusan proyek dibatalkan.")

    def update_project_info(self):
        if self.current_project_id:
            projects = self.db.get_projects(self.user_id)
            project = next((p for p in projects if p[0] == self.current_project_id), None)
            if project:
                self.current_project_name = project[1]
                self.total_project = self.parse_currency(project[6])
                
                self.project_name_label.setText(f"Proyek: {self.current_project_name}")
                self.sales_label.setText(f"<b>Sales:</b> {project[2]}")
                self.worker_label.setText(f"<b>Tukang:</b> {project[3]}")
                self.date_label.setText(f"<b>Tanggal:</b> {project[4]} - {project[5]}")
                self.total_label.setText(f"<b>   Total:</b> {self.format_currency(self.total_project)}")
                self.dp_label.setText(f"<b>DP:</b> {self.format_currency(self.parse_currency(project[7]))}")
                
                # Enable rich text interpretation for labels
                for label in [self.sales_label, self.worker_label, self.date_label, self.total_label, self.dp_label]:
                    label.setTextFormat(Qt.RichText)
            else:
                self.reset_project_info()
        else:
            self.reset_project_info()
        
        self.update_total_profit()

    def reset_project_info(self):
        self.project_name_label.setText("Belum ada proyek dipilih")
        self.sales_label.setText("")
        self.worker_label.setText("")
        self.date_label.setText("")
        self.total_label.setText("")
        self.dp_label.setText("")
        self.current_project_id = None
        self.current_project_name = ""
        self.total_project = 0

    def open_edit_dialog(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin diedit terlebih dahulu.")
            return

        selected_row = selected_items[0].row()
        dialog = AddMaterialDialog(self)
        
        initial_data = []
        for col in range(self.table.columnCount()):
            item = self.table.item(selected_row, col)
            if item is not None:
                if col in [3, 4]:  # Harga Satuan and Total columns
                    # Parse currency for editing
                    value = self.parse_currency(item.text())
                    initial_data.append(str(value))
                else:
                    initial_data.append(item.text())
            else:
                initial_data.append("")
        
        dialog.load_data(initial_data)

        if dialog.exec_():
            if dialog.validate_data():
                data = dialog.get_data()
                for col, item_text in enumerate(data):
                    if col in [3, 4]:  # Harga Satuan and Total columns
                        formatted_price = self.format_currency(float(item_text))
                        self.table.setItem(selected_row, col, QTableWidgetItem(formatted_price))
                    else:
                        self.table.setItem(selected_row, col, QTableWidgetItem(str(item_text) if item_text is not None else ""))
                
                material_id = self.db.get_material_usage(self.current_project_id, self.user_id)[selected_row][0]
                self.db.update_material_usage(material_id, data, self.user_id)
                self.update_total_price()
            else:
                QMessageBox.warning(self, "Input Tidak Valid", "Quantity dan Harga Satuan harus berupa angka.")

    def delete_selected_row(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Peringatan", "Silakan pilih baris yang ingin dihapus terlebih dahulu.")
            return

        reply = QMessageBox.question(self, 'Konfirmasi Hapus', 'Anda yakin ingin menghapus data ini?',
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            selected_row = selected_items[0].row()

            materials = self.db.get_material_usage(self.current_project_id, self.user_id)
            if selected_row < len(materials):
                record_id = materials[selected_row][0]
                self.db.delete_record(self.table_name, record_id, self.user_id)
            else:
                QMessageBox.warning(self, "Error", "Tidak dapat menemukan data yang akan dihapus.")
                return

            self.table.removeRow(selected_row)
            self.update_total_price()
            QMessageBox.information(self, "Hapus Data", "Data berhasil dihapus.")

    def update_total_price(self):
        total_price = 0
        for row in range(self.table.rowCount()):
            total_item = self.table.item(row, 4)  # Total column
            if total_item and total_item.text():
                try:
                    price = self.parse_currency(total_item.text())
                    total_price += price
                except ValueError:
                    pass  # Skip if the value can't be converted to float
        self.total_price_label.setText(f"Total Harga: {self.format_currency(total_price)}")
        self.update_total_profit()

    def update_total_profit(self):
        total_price = self.parse_currency(self.total_price_label.text().split(": ")[1])
        total_profit = self.total_project - total_price
        self.total_profit_label.setText(f"Total Keuntungan: {self.format_currency(total_profit)}")

    def get_add_dialog(self):
        return AddMaterialDialog(self)

    def open_add_dialog(self):
        if self.current_project_id:
            dialog = self.get_add_dialog()
            if dialog.exec_():
                if dialog.validate_data():
                    data = dialog.get_data()
                    self.db.insert_material_usage(self.current_project_id, data, self.user_id)
                    self.add_row(data)
                    self.update_total_price()
                else:
                    QMessageBox.warning(self, "Input Tidak Valid", "Quantity dan Harga Satuan harus berupa angka.")
        else:
            QMessageBox.warning(self, "No Project Selected", "Please select or create a project first.")

    def add_row(self, data):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        for column, item in enumerate(data):
            if column in [3, 4]:  # Harga Satuan and Total columns
                formatted_price = format_rupiah(item)
                self.table.setItem(row_position, column, QTableWidgetItem(formatted_price))
            else:
                self.table.setItem(row_position, column, QTableWidgetItem(str(item)))
        self.update_total_price()
    
    def format_price(self, price):
        try:
            price_float = float(price)
            return f"Rp {price_float:,.0f}".replace(',', '.')
        except ValueError:
            return price 
    
    def add_additional_info(self, sheet):
        project = next((p for p in self.db.get_projects(self.user_id) if p[0] == self.current_project_id), None)
        if project:
            project_details = [
                ("Nama Proyek", project[1]),
                ("Sales", project[2]),
                ("Tukang", project[3]),
                ("Tanggal Mulai", project[4]),
                ("Tanggal Selesai", project[5]),
                ("Total Proyek", self.format_currency(self.parse_currency(project[6]))),
               ("DP", self.format_currency(self.parse_currency(project[7])))
            ]
            for label, value in project_details:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
                sheet.cell(row=row, column=2, value=value)

        total_price = self.parse_currency(self.total_price_label.text().split(": ")[1])
        total_profit = self.parse_currency(self.total_profit_label.text().split(": ")[1])

        sheet.append([""] * sheet.max_column)
    
        sheet.cell(row=sheet.max_row + 1, column=1, value="Total Harga").font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=2, value=self.format_currency(total_price))
    
        sheet.cell(row=sheet.max_row + 1, column=1, value="Total Keuntungan").font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=2, value=self.format_currency(total_profit))

    def close_book(self):
        pass

    def view_history(self):
        pass

class DateInputDialog(QDialog):
    def __init__(self, parent=None, title="Select Date"):
        super().__init__(parent)
        setup_error_handling()
        self.setWindowTitle(title)
        self.layout = QVBoxLayout(self)
        
        self.date_edit = QDateEdit(QDate.currentDate(), self)
        self.date_edit.setCalendarPopup(True)
        self.layout.addWidget(QLabel(title))
        self.layout.addWidget(self.date_edit)
        
        self.ok_button = QPushButton("OK", self)
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

    def get_date(self):
        return self.date_edit.date()

class PhotoViewerDialog(QDialog):
    def __init__(self, parent=None, photo_path=None):
        super().__init__(parent)
        self.setWindowTitle("Lihat Foto")
        self.photo_path = photo_path
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        photo_label = QLabel(self)
        if self.photo_path and os.path.exists(self.photo_path):
            pixmap = QPixmap(self.photo_path)
            if not pixmap.isNull():
                photo_label.setPixmap(pixmap.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                photo_label.setText("Error loading image")
        else:
            photo_label.setText("Image not found or invalid path")
        layout.addWidget(photo_label)
        
        button_layout = QHBoxLayout()
        
        close_button = QPushButton("Tutup", self)
        close_button.clicked.connect(self.close)
        button_layout.addWidget(close_button)
        
        open_file_button = QPushButton("Buka File", self)
        open_file_button.clicked.connect(self.open_file)
        button_layout.addWidget(open_file_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)

    def open_file(self):
        if os.path.exists(self.photo_path):
            os.startfile(self.photo_path)